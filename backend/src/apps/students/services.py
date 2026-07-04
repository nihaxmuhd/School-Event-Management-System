from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO

from django.contrib.auth.models import User
from django.core.exceptions import ValidationError
from django.db import transaction
from django.db.models import Q
try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - optional dependency in this workspace
    load_workbook = None

from .models import Student
from .validators import validate_division, validate_gender, validate_student_class
from apps.houses.models import House


@dataclass
class ImportPreviewRow:
    row_number: int
    admission_no: str
    student_name: str
    gender: str
    student_class: str
    division: str
    house_id: str
    is_valid: bool
    errors: list[str]


class StudentService:
    @staticmethod
    def create_student(*, data: dict, user: User | None = None) -> Student:
        student = Student(**data)
        student.created_by = user
        student.updated_by = user
        student.full_clean()
        student.save()
        return student

    @staticmethod
    def update_student(*, student: Student, data: dict, user: User | None = None) -> Student:
        for key, value in data.items():
            setattr(student, key, value)
        student.updated_by = user
        student.full_clean()
        student.save()
        return student

    @staticmethod
    def soft_delete_student(*, student: Student, user: User | None = None) -> Student:
        student.is_deleted = True
        student.status = Student.StatusChoices.INACTIVE
        student.updated_by = user
        student.save(update_fields=['is_deleted', 'status', 'updated_by', 'updated_at'])
        return student

    @staticmethod
    def duplicate_admission_exists(admission_no: str, exclude_id=None) -> bool:
        queryset = Student.objects.filter(admission_no__iexact=admission_no, is_deleted=False)
        if exclude_id:
            queryset = queryset.exclude(id=exclude_id)
        return queryset.exists()

    @staticmethod
    def search(queryset, search: str):
        if not search:
            return queryset
        return queryset.filter(Q(admission_no__icontains=search) | Q(student_name__icontains=search))

    @staticmethod
    def parse_excel(file_bytes: bytes) -> list[list[str]]:
        if load_workbook is None:
            raise RuntimeError('openpyxl is required for Excel import support.')
        workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
        worksheet = workbook.active
        rows = list(worksheet.iter_rows(values_only=True))
        return [list(row) for row in rows]

    @classmethod
    def preview_import(cls, file_bytes: bytes, existing_admission_nos: set[str] | None = None) -> list[ImportPreviewRow]:
        rows = cls.parse_excel(file_bytes)
        existing_admission_nos = existing_admission_nos or set()
        house_codes = set(House.objects.filter(is_active=True).values_list('code', flat=True))
        preview_rows: list[ImportPreviewRow] = []
        seen_in_file: set[str] = set()
        for index, row in enumerate(rows[1:], start=2):
            admission_no = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ''
            student_name = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
            student_class = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ''
            division = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ''
            gender = str(row[4]).strip() if len(row) > 4 and row[4] is not None else ''
            house_id = str(row[5]).strip() if len(row) > 5 and row[5] is not None else ''
            errors = []

            if not admission_no:
                errors.append('Admission number is required.')
            elif admission_no in seen_in_file:
                errors.append('Duplicate admission number in the uploaded file.')
            elif admission_no in existing_admission_nos:
                errors.append('Admission number already exists in the database.')
            seen_in_file.add(admission_no)

            if not student_name:
                errors.append('Student name is required.')
            try:
                validate_student_class(student_class)
            except Exception as exc:
                errors.append(str(exc))
            try:
                validate_division(division)
            except Exception as exc:
                errors.append(str(exc))
            try:
                validate_gender(gender)
            except Exception as exc:
                errors.append(str(exc))
            if house_id and house_id not in house_codes:
                errors.append('House code does not exist or is inactive.')

            preview_rows.append(ImportPreviewRow(
                row_number=index,
                admission_no=admission_no,
                student_name=student_name,
                gender=gender,
                student_class=student_class,
                division=division,
                house_id=house_id,
                is_valid=not errors,
                errors=errors,
            ))
        return preview_rows

    @classmethod
    @transaction.atomic
    def import_students(cls, preview_rows: list[ImportPreviewRow], user: User | None = None):
        created_students = []
        for row in preview_rows:
            if not row.is_valid:
                continue
            house = House.objects.filter(code=row.house_id, is_active=True).first()
            if house is None:
                raise ValidationError(f'House code {row.house_id} is invalid.')
            created_students.append(Student.objects.create(
                admission_no=row.admission_no,
                student_name=row.student_name,
                gender=row.gender,
                student_class=row.student_class,
                division=row.division,
                house=house,
                team_leader_assigned=False,
                status=Student.StatusChoices.ACTIVE,
                created_by=user,
                updated_by=user,
            ))
        return created_students
